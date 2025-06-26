#from reportlab.lib.pagesizes import A4
#from reportlab.pdfgen import canvas
#from reportlab.lib.units import cm
import logging
logger = logging.getLogger(__name__)
from .models import ElectionConfig
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Voter, Election, Candidate, Vote, Admin, Poste, Candidat, VoteParPoste, EffectifClasse, ResponsableClassePoste, ResponsableClasseCandidat, VoteResponsableClasse, Club, VoteClub, ClubCandidat, Configuration
import pandas as pd
from django.http import HttpResponse
from django.template.loader import render_to_string
import io, random, string
from django.contrib.auth.decorators import login_required
from django.db import models
from .utils import matricule_required, admin_required, check_access
from django.core.serializers import serialize
import json
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import json
from .forms import VoteForm, EffectifClasseForm, CandidatureForm, ClubCandidatureForm, ResponsableClasseCandidatureForm
from .forms import UploadCandidatExcelForm, UploadEmailsExcelForm
import openpyxl
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import Count, F, Prefetch
from django.core.mail import send_mail
from django.core.files.storage import default_storage
import datetime
from datetime import datetime







def home(request):
    return render(request, 'vote/accueil.html')

def vote_classe(request):
    return render(request, 'vote/vote_classe.html')




from django.shortcuts import render
from .models import Poste, VoteParPoste

from django.shortcuts import render
from .models import Poste, Candidat, VoteParPoste

def resultats_view(request):
    resultats = []
    chart_data = []
    classement = []

    postes = Poste.objects.all()

    for poste in postes:
        candidats = poste.candidats.all()  # ou Candidat.objects.filter(poste=poste)
        candidat_votes = []
        labels = []
        votes = []
        classement_par_poste = []
        total_votes = 0

        for candidat in candidats:
            total = VoteParPoste.objects.filter(candidat=candidat).count()
            candidat_votes.append((candidat, total))
            labels.append(candidat.nom)
            votes.append(total)
            classement_par_poste.append({
                "name": candidat.nom,
                "votes": total,
            })
            total_votes += total

        # Calcul des pourcentages + tri du classement
        for c in classement_par_poste:
            c["pourcentage"] = round((c["votes"] / total_votes) * 100, 1) if total_votes > 0 else 0.0

        classement_par_poste.sort(key=lambda x: x['votes'], reverse=True)

        # Remplir les structures
        resultats.append((poste, candidat_votes))
        chart_data.append({
            'poste': poste.nom,
            'labels': labels,
            'votes': votes
        })
        classement.append({
            "poste": poste.nom,
            "candidats": classement_par_poste,
            "total_votes": total_votes
        })

    return render(request, 'vote/resultats.html', {
        'resultats': resultats,
        'chart_data': chart_data,
        'classement': classement
    })






from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Voter

def login_view(request):
    if request.method == 'POST':
        matricule = request.POST.get('matricule', '').strip()

        # 1. D'abord vérifier dans les nouveaux admins (modèle Admin)
        try:
            admin = Admin.objects.get(matricule=matricule, est_actif=True)
            request.session['admin_id'] = admin.id
            request.session['admin_matricule'] = admin.matricule
            return redirect('admin_dashboard')
        except Admin.DoesNotExist:
            pass  # Ce n'est pas un nouveau admin

        # 2. Ensuite vérifier dans les anciens admins (modèle Voter)
        try:
            voter = Voter.objects.get(matricule=matricule)
            if voter.is_admin:  # Si c'est un ancien admin
                request.session['voter_id'] = voter.id
                request.session['admin_matricule'] = voter.matricule
                return redirect('admin_dashboard')
            
            # Sinon c'est un votant normal
            request.session['voter_id'] = voter.id
            return redirect('vote')
        except Voter.DoesNotExist:
            messages.error(request, "Matricule incorrect. Veuillez réessayer.")

    return render(request, 'vote/login.html')


def logout_view(request):
    request.session.flush()  # Supprime toute la session
    return redirect('login')


@matricule_required
def vote_view(request):

    config = Configuration.get_config()
    if not config.votes_ouverts:
        return redirect('access_closed_vote')
    voter_id = request.session.get('voter_id')
    if not voter_id:
        return redirect('login')

    voter = Voter.objects.get(id=voter_id)
    if voter.has_voted:
        return render(request, 'vote/already_voted.html')

    election = Election.objects.filter(is_active=True).first()
    candidates = Candidate.objects.filter(election=election)

    if request.method == 'POST':
        candidate_id = request.POST.get('candidate')
        candidate = Candidate.objects.get(id=candidate_id)
        Vote.objects.create(voter=voter, candidate=candidate)
        voter.has_voted = True
        voter.save()
        return redirect('thank_you')

    return render(request, 'vote/vote.html', {'candidates': candidates})

def thank_you(request):
    return render(request, 'vote/thank_you.html')

"""
def generate_report(request):
    # Collecte des données
    total_votes = Vote.objects.count()
    candidates = Candidate.objects.all()
    data = []
    for c in candidates:
        count = Vote.objects.filter(candidate=c).count()
        data.append({'Candidat': c.name, 'Votes': count})

    df = pd.DataFrame(data)
    format = request.GET.get("format", "pdf")

    if format == "excel":
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rapport_election.xlsx"'
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False)
        return response

    elif format == "pdf":
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        # Titre
        p.setFont("Helvetica-Bold", 16)
        p.drawString(3 * cm, 27 * cm, "Rapport des élections")
        p.setFont("Helvetica", 12)

        # Total
        p.drawString(3 * cm, 25.5 * cm, f"Total de votes : {total_votes}")

        # Tableau des résultats
        y = 24 * cm
        p.drawString(3 * cm, y, "Candidat")
        p.drawString(10 * cm, y, "Nombre de votes")
        y -= 1 * cm

        for row in data:
            p.drawString(3 * cm, y, row['Candidat'])
            p.drawString(10 * cm, y, str(row['Votes']))
            y -= 0.8 * cm
            if y < 2 * cm:
                p.showPage()
                y = 27 * cm

        p.showPage()
        p.save()

        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    else:
        return HttpResponse("Format non supporté", status=400)
"""
        
@matricule_required
def vote_view(request):

    config = Configuration.get_config()
    if not config.votes_ouverts:
        return redirect('access_closed_vote')
    voter_id = request.session.get('voter_id')
    if not voter_id:
        return redirect('login')

    voter = Voter.objects.get(id=voter_id)
    if voter.has_voted:
        return render(request, 'vote/already_voted.html')

    postes = Poste.objects.prefetch_related('candidats').all()

    if request.method == 'POST':
        form = VoteForm(request.POST, postes=postes)
        if form.is_valid():
            for poste in postes:
                champ = f"poste_{poste.id}"
                candidat_id = form.cleaned_data.get(champ)
                if candidat_id:
                    candidat = Candidat.objects.get(id=candidat_id)
                    VoteParPoste.objects.create(
                        voter=voter,
                        candidat=candidat,
                        poste=poste
                    )
            voter.has_voted = True
            voter.save()
            return redirect('merci')
    else:
        form = VoteForm(postes=postes)

    return render(request, 'vote/vote.html', {
        'form': form,
        'postes': postes
    })


def accueil(request):
    return render(request, 'accueil.html')


def merci(request):
    return render(request, 'vote/merci.html')

def liste_candidats(request):
    candidats = Candidat.objects.select_related('poste').all()
    return render(request, 'vote/liste_candidats.html', {'candidats': candidats})


def upload_candidats(request):
    if request.method == 'POST':
        form = UploadCandidatExcelForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier_excel']
            try:
                wb = openpyxl.load_workbook(fichier)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):  # On saute la ligne d'en-tête
                    nom, prenom, poste_nom = row
                    if nom and prenom and poste_nom:
                        poste, _ = Poste.objects.get_or_create(nom=poste_nom)
                        Candidat.objects.create(nom=nom, prenom=prenom, poste=poste)
                messages.success(request, "Importation réussie.")
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {e}")
    else:
        form = UploadCandidatExcelForm()

    return render(request, 'vote/upload_candidats.html', {'form': form})




def vote_aes(request):
    # Récupère tous les postes avec leurs candidats en une seule requête optimisée
    postes = Poste.objects.prefetch_related('candidats').order_by('code')
    
    # Préparer les données pour le template
    postes_data = []
    for poste in postes:
        candidats_data = []
        for candidat in poste.candidats.all():
            # Formatage des données du candidat
            nom_complet = f"{candidat.nom}"
            if candidat.classe:
                nom_complet += f" ({candidat.classe})"
            
            candidats_data.append({
                'id': candidat.id,
                'nom': candidat.nom,
                'nom_complet': nom_complet,
                'photo': candidat.photo.url if candidat.photo else None
            })
        
        postes_data.append({
            'poste': poste,
            'candidats': candidats_data
        })
    
    return render(request, 'vote/vote_aes.html', {
        'postes': postes_data,
        'page_title': 'Vote pour l\'Amicale des Élèves'
    })


@require_POST
def voter(request):
    try:
        data = json.loads(request.body)
        poste_id = data.get("poste_id")
        candidat_id = data.get("candidat_id")

        voter_id = request.session.get("voter_id")
        if not voter_id:
            return JsonResponse({"error": "Non authentifié"}, status=403)

        voter = Voter.objects.get(id=voter_id)

        # Vérifie si la personne a déjà voté pour ce poste
        if VoteParPoste.objects.filter(voter=voter, poste_id=poste_id).exists():
            return JsonResponse({"error": "Déjà voté pour ce poste"}, status=400)

        candidat = Candidat.objects.get(id=candidat_id, poste_id=poste_id)

        VoteParPoste.objects.create(voter=voter, candidat=candidat, poste_id=poste_id)

        # Optionnel : si tous les postes sont remplis, on peut marquer le voter comme ayant voté globalement
        total_postes = Poste.objects.count()
        votes_effectues = VoteParPoste.objects.filter(voter=voter).count()
        if votes_effectues == total_postes:
            voter.has_voted = True
            voter.save()

        # Recalculer les résultats pour ce poste
        candidats = Candidat.objects.filter(poste_id=poste_id)
        total_votes = VoteParPoste.objects.filter(poste_id=poste_id).count()

        results = {}
        for c in candidats:
            votes = VoteParPoste.objects.filter(candidat=c).count()
            pourcentage = round((votes / total_votes) * 100, 2) if total_votes else 0
            results[c.id] = pourcentage

        return JsonResponse({"success": True, "results": results})
    
    except (KeyError, json.JSONDecodeError):
        return JsonResponse({"error": "Requête invalide"}, status=400)
    except Candidat.DoesNotExist:
        return JsonResponse({"error": "Candidat introuvable"}, status=404)
    


@matricule_required
def statistiques_view(request):
    # Récupération des votants, votes et postes
    votants = Voter.objects.all()
    votes = VoteParPoste.objects.select_related('voter', 'candidat', 'poste')
    postes = Poste.objects.all()

    # Listes des classes et filières selon les choix définis dans le modèle Voter
    classes = [c[0] for c in Voter.CLASS_CHOICES]
    filieres = [f[0] for f in Voter.FILIERE_CHOICES]

    # Comptages globaux de votants par classe et par filière
    classe_counts = votants.values('classe').annotate(total=Count('id')).order_by('-total')
    filiere_counts = votants.values('filiere').annotate(total=Count('id')).order_by('-total')

    # Initialisation des structures pour stocker les stats
    stats_par_classe = {}
    stats_par_filiere = {}
    votes_par_classe = {}
    votes_par_filiere = {}

    for poste in postes:
        poste_nom = poste.nom

        # Stats votes par classe pour ce poste
        stats_par_classe[poste_nom] = {}
        for classe in classes:
            stats = (
                votes.filter(voter__classe=classe, poste=poste)
                .values('candidat__nom')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            stats_par_classe[poste_nom][classe] = {
                "labels": [s['candidat__nom'] for s in stats],
                "votes": [s['total'] for s in stats]
            }

        # Stats votes par filière pour ce poste
        stats_par_filiere[poste_nom] = {}
        for filiere in filieres:
            stats = (
                votes.filter(voter__filiere=filiere, poste=poste)
                .values('candidat__nom')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            stats_par_filiere[poste_nom][filiere] = {
                "labels": [s['candidat__nom'] for s in stats],
                "votes": [s['total'] for s in stats]
            }

        # Votes détaillés par classe pour ce poste (liste simple)
        for classe in classes:
            votes_list = (
                votes.filter(voter__classe=classe, poste=poste)
                .values('poste__nom', 'candidat__nom')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            votes_par_classe.setdefault(classe, []).extend(votes_list)

        # Votes détaillés par filière pour ce poste (liste simple)
        for filiere in filieres:
            votes_list = (
                votes.filter(voter__filiere=filiere, poste=poste)
                .values('poste__nom', 'candidat__nom')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            votes_par_filiere.setdefault(filiere, []).extend(votes_list)

    # Debug console (optionnel, à commenter en production)
    print(f"Total votants: {votants.count()}")
    print(f"Total postes: {postes.count()}")
    print(f"Total votes: {votes.count()}")

    return render(request, 'vote/statistiques.html', {
        'classe_counts': classe_counts,
        'filiere_counts': filiere_counts,
        'stats_par_classe': json.dumps(stats_par_classe),
        'stats_par_filiere': json.dumps(stats_par_filiere),
        'votes_par_classe': votes_par_classe,
        'votes_par_filiere': votes_par_filiere,
        'classes': classes,
        'filieres': filieres,
        'postes': postes,
    })

"""
#################VOTE RESPONSABLES DE CLASSE################

def vote_par_classe_view(request):
    postes = Poste.objects.filter(nom__icontains="responsable de classe")

    if request.method == "POST":
        matricule = request.POST.get("matricule")
        filiere = request.POST.get("filiere")
        classe = request.POST.get("classe")
        candidat_id = request.POST.get("candidat_id")

        try:
            voter = Voter.objects.get(matricule=matricule)
        except Voter.DoesNotExist:
            messages.error(request, "Matricule introuvable.")
            return redirect("vote_par_classe")

        if voter.classe != classe or voter.filiere != filiere:
            messages.error(request, "Le matricule ne correspond pas à la classe ou filière sélectionnée.")
            return redirect("vote_par_classe")

        if VoteParPoste.objects.filter(voter=voter, poste__nom__icontains="responsable de classe").exists():
            messages.warning(request, "Vous avez déjà voté pour ce poste.")
            return redirect("vote_par_classe")

        candidat = get_object_or_404(Candidat, id=candidat_id)
        poste = candidat.poste

        VoteParPoste.objects.create(voter=voter, candidat=candidat, poste=poste)
        voter.has_voted = True
        voter.save()

        messages.success(request, f"Vote enregistré pour {candidat.nom}")
        return redirect("vote_par_classe")

    filieres = Voter.FILIERE_CHOICES
    classes_par_filiere = {}
    for classe, _ in Voter.CLASS_CHOICES:
        filiere_associee = Voter.CLASS_TO_FILIERE[classe]
        classes_par_filiere.setdefault(filiere_associee, []).append(classe)

    context = {
        "filieres": filieres,
        "classes_par_filiere": classes_par_filiere,
        "postes": postes,
    }
    return render(request, "vote_classe.html", context)
"""




#################GESTION DES MATRICULES################


def generer_codes_view(request):
    if request.method == "POST":
        form = EffectifClasseForm(request.POST)

        if form.is_valid():
            classe = form.cleaned_data["classe"]
            nombre_eleves = form.cleaned_data["nombre_eleves"]

            # Récupérer ou créer l'instance
            instance, created = EffectifClasse.objects.get_or_create(
                classe=classe,
                defaults={'nombre_eleves': nombre_eleves}
            )
            
            # Si l'instance existait déjà, mettre à jour le nombre d'élèves
            if not created:
                instance.nombre_eleves = nombre_eleves
                instance.codes_genere = False  # Réinitialiser ce flag
                instance.save()

            # Générer les codes
            instance.generer_codes()

            messages.success(request, f"{instance.nombre_eleves} codes générés pour {instance.classe}.")
            return redirect("generer_codes")
    else:
        form = EffectifClasseForm()

    context = {
        "form": form,
        "classes_existantes": EffectifClasse.objects.all()
    }
    return render(request, "vote/generer_codes.html", context)



def reset_affichage_codes(request, classe_id):
    if request.method == "POST":
        classe = get_object_or_404(EffectifClasse, id=classe_id)
        classe.codes_genere = False
        classe.save()
        messages.info(request, f"L'état de la classe {classe.classe} a été réinitialisé (affichage uniquement).")
    return redirect("generer_codes")



from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404

@require_POST
def reset_codes_view(request, classe_id):
    effectif = get_object_or_404(EffectifClasse, id=classe_id)
    Voter.objects.filter(classe=effectif.classe).delete()
    effectif.codes_genere = False
    effectif.save()
    messages.success(request, f"Les codes pour la classe {effectif.classe} ont été réinitialisés.")
    return redirect("generer_codes")

"""
################ENREGISTREMENT DES CODES##################


def generer_codes_view(request):
    context = {}
    if request.method == "POST":
        effectif = int(request.POST.get("effectif"))
        classe = request.POST.get("classe")
        filiere = request.POST.get("filiere")

        codes_gen = []

        for _ in range(effectif):
            code = ''.join(random.choices(string.digits, k=4)) + ''.join(random.choices(string.ascii_uppercase, k=3))
            while CodeVote.objects.filter(code=code).exists():
                code = ''.join(random.choices(string.digits, k=4)) + ''.join(random.choices(string.ascii_uppercase, k=3))

            # Enregistrement en base
            CodeVote.objects.create(code=code, classe=classe, filiere=filiere)
            codes_gen.append(code)

        context["codes"] = codes_gen
        context["effectif"] = effectif
        context["classe"] = classe
        context["filiere"] = filiere

    return render(request, "vote/generer_codes.html", context)
"""


def export_voters_excel(request):
    # Création d’un classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matricules"

    # En-têtes
    ws.append(["Matricule", "Classe"])

    # Lignes des données
    for voter in Voter.objects.all():
        ws.append([voter.matricule, voter.classe])

    # Réponse HTTP avec fichier attaché
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="matricules.xlsx"'
    wb.save(response)
    return response


def telechargement_view(request):
    return render(request, "vote/telechargement.html")



################GESTION DES CANDIDATURES##################


@matricule_required
def candidature_view(request):
    voter = request.voter  # injecté par le décorateur

    if request.method == 'POST':
        form = CandidatureForm(request.POST, request.FILES)
        if form.is_valid():
            poste = form.cleaned_data['poste']
            prenom = form.cleaned_data['prenom']
            nom = form.cleaned_data['nom']
            classe_form = form.cleaned_data['classe']
            photo = form.cleaned_data['photo']

            if classe_form != voter.classe:
                messages.error(request, "La classe choisie ne correspond pas à votre matricule.")
                return render(request, 'vote/candidature.html', {'form': form})

            nom_complet = f"{prenom} {nom}"
            Candidat.objects.create(
                nom=nom_complet,
                poste=poste,
                photo=photo,
                classe=classe_form,
                filiere=voter.filiere
            )
            messages.success(request, "Votre candidature a été enregistrée avec succès.")
            return redirect('accueil')
    else:
        form = CandidatureForm()

    return render(request, 'vote/candidature.html', {'form': form})


################GESTION DES MAILS##################




def contact_email(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        message = request.POST.get('message')

        send_mail(
            subject="Message reçu via la plateforme de vote",
            message=f"Expéditeur : {email}\n\nMessage :\n{message}",
            from_email="ton_adresse@example.com",
            recipient_list=["landrysanam1@gmail.com"],
            fail_silently=False,
        )

        messages.success(request, "Votre message a bien été envoyé !")
        return redirect('accueil')  # ou une autre page

    return redirect('accueil')


################CONNEXION ADMIN##################

@admin_required
def admin_dashboard_view(request):
    # Récupération du type de vote sélectionné (par défaut: AES)
    vote_type = request.GET.get('vote_type', 'aes')
    
    # Données communes à tous les types de votes
    common_stats = {
        'total_votants': Voter.objects.count(),
        'participations': [],
        'votes_exprimes': 0  # Initialisation ajoutée ici
    }
    
    # Calcul des participations par classe (commun à tous les types)
    for classe in Voter.CLASS_CHOICES:
        classe_nom = classe[0]
        inscrits = Voter.objects.filter(classe=classe_nom).count()
        
        # Votes exprimés selon le type sélectionné
        if vote_type == 'aes':
            votants = VoteParPoste.objects.filter(voter__classe=classe_nom)\
                         .values('voter').distinct().count()
        elif vote_type == 'club':
            votants = VoteClub.objects.filter(voter__classe=classe_nom)\
                         .values('voter').distinct().count()
        elif vote_type == 'classe':
            votants = VoteResponsableClasse.objects.filter(voter__classe=classe_nom)\
                         .values('voter').distinct().count()
        
        taux = 0
        if inscrits > 0:
            taux = min(100, round((votants / inscrits) * 100))
        
        common_stats['participations'].append({
            'classe': classe_nom,
            'taux': taux,
            'inscrits': inscrits,
            'votants': votants
        })
    
    # Calcul des votes exprimés globaux
    if vote_type == 'aes':
        common_stats['votes_exprimes'] = VoteParPoste.objects.values('voter').distinct().count()
    elif vote_type == 'club':
        common_stats['votes_exprimes'] = VoteClub.objects.values('voter').distinct().count()
    elif vote_type == 'classe':
        common_stats['votes_exprimes'] = VoteResponsableClasse.objects.values('voter').distinct().count()
    
    # Préparation du contexte
    context = {
        'stats': common_stats,
        'postes': Poste.objects.all().order_by('code'),
        'clubs': Club.objects.all(),
        'postes_responsables': ResponsableClassePoste.objects.all().order_by('ordre'),
        'classes': Voter.CLASS_CHOICES,
        'current_vote_type': vote_type
    }
    
    
    return render(request, 'admin.html', context)


################AJOUT ELECTEUR##################


@require_POST
@admin_required
def ajouter_electeur(request):
    matricule = request.POST.get('matricule')
    classe = request.POST.get('classe')

    if not matricule or not classe:
        return JsonResponse({'success': False, 'message': "Champs manquants."})

    if Voter.objects.filter(matricule=matricule).exists():
        return JsonResponse({'success': False, 'message': "Matricule déjà utilisé."})

    Voter.objects.create(matricule=matricule, classe=classe)
    return JsonResponse({'success': True, 'message': "Électeur ajouté avec succès."})


################SUPPRESSION ELECTEUR##################


@require_POST
@admin_required
def supprimer_electeur(request):
    matricule = request.POST.get('matricule')
    try:
        voter = Voter.objects.get(matricule=matricule)
        voter.delete()
        return JsonResponse({'success': True, 'message': "Électeur supprimé."})
    except Voter.DoesNotExist:
        return JsonResponse({'success': False, 'message': "Aucun électeur trouvé."})



################AJOUT CANDIDATS##################



@require_POST
@admin_required
def ajouter_candidat(request):
    type_candidat = request.POST.get('type')
    nom = request.POST.get('nom')
    photo = request.FILES.get('photo')
    
    if not nom:
        return JsonResponse({'success': False, 'message': "Le nom est requis."})

    try:
        if type_candidat == 'aes':
            poste_id = request.POST.get('poste')
            classe = request.POST.get('classe')
            poste = Poste.objects.get(id=poste_id)
            candidat = Candidat.objects.create(
                nom=nom, 
                poste=poste, 
                classe=classe, 
                photo=photo
            )
            
        elif type_candidat == 'club':
            club_id = request.POST.get('club')
            classe = request.POST.get('classe')
            slogan = request.POST.get('slogan', '')
            programme = request.POST.get('programme', '')
            club = Club.objects.get(id=club_id)
            candidat = ClubCandidat.objects.create(
                nom=nom,
                club=club,
                classe=classe,
                photo=photo,
                slogan=slogan,
                programme=programme
            )
            
        elif type_candidat == 'classe':
            poste_id = request.POST.get('poste')
            classe = request.POST.get('classe')
            programme = request.POST.get('programme', '')
            poste = ResponsableClassePoste.objects.get(id=poste_id)
            candidat = ResponsableClasseCandidat.objects.create(
                nom=nom,
                poste=poste,
                classe=classe,
                programme=programme,
                photo=photo
            )
            
        else:
            return JsonResponse({'success': False, 'message': "Type de candidat invalide."})

        return JsonResponse({'success': True, 'message': f"{nom} ajouté avec succès."})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"Erreur: {str(e)}"})


@require_POST
@admin_required
def supprimer_candidat(request):
    type_candidat = request.POST.get('type')
    nom = request.POST.get('nom')
    matricule = request.POST.get('matricule', None)

    try:
        if type_candidat == 'aes':
            candidat = Candidat.objects.get(nom=nom)
        elif type_candidat == 'club':
            if matricule:
                candidat = ClubCandidat.objects.get(nom=nom, matricule=matricule)
            else:
                candidat = ClubCandidat.objects.get(nom=nom)
        elif type_candidat == 'classe':
            if matricule:
                candidat = ResponsableClasseCandidat.objects.get(nom=nom, matricule=matricule)
            else:
                candidat = ResponsableClasseCandidat.objects.get(nom=nom)
        else:
            return JsonResponse({'success': False, 'message': "Type de candidat invalide."})

        candidat.delete()
        return JsonResponse({'success': True, 'message': "Candidat supprimé avec succès."})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"Erreur: {str(e)}"})




################OUVERTURE-FERMETURE/VOTES-CANDIDATURES##################


@require_POST
@admin_required
def controler_periode(request):
    logger.info(f"Requête reçue pour controler_periode - Utilisateur: {request.user} - Données: {request.POST}")
    
    try:
        config = Configuration.get_config()
        action = request.POST.get('action')
        
        if not action:
            logger.warning("Action non spécifiée dans la requête")
            return JsonResponse({'success': False, 'message': 'Action non spécifiée'}, status=400)
        
        logger.info(f"Traitement de l'action: {action} - Configuration actuelle: "
                   f"Candidatures: {config.candidatures_ouvertes}, Votes: {config.votes_ouverts}")
        
        if action == 'open_candidatures':
            if config.candidatures_ouvertes:
                logger.warning("Tentative d'ouverture des candidatures déjà ouvertes")
                return JsonResponse({'success': False, 'message': 'Les candidatures sont déjà ouvertes!'})
            
            config.candidatures_ouvertes = True
            config.save()
            logger.info("Candidatures ouvertes avec succès")
            return JsonResponse({'success': True, 'message': 'Candidatures ouvertes avec succès!'})
        
        elif action == 'close_candidatures':
            if not config.candidatures_ouvertes:
                logger.warning("Tentative de fermeture des candidatures déjà fermées")
                return JsonResponse({'success': False, 'message': 'Les candidatures sont déjà fermées!'})
            
            config.candidatures_ouvertes = False
            config.save()
            logger.info("Candidatures fermées avec succès")
            return JsonResponse({'success': True, 'message': 'Candidatures fermées avec succès!'})
        
        elif action == 'open_votes':
            if config.votes_ouverts:
                logger.warning("Tentative d'ouverture des votes déjà ouverts")
                return JsonResponse({'success': False, 'message': 'Les votes sont déjà ouverts!'})
            
            config.votes_ouverts = True
            config.save()
            logger.info("Votes ouverts avec succès")
            return JsonResponse({'success': True, 'message': 'Votes ouverts avec succès!'})
        
        elif action == 'close_votes':
            if not config.votes_ouverts:
                logger.warning("Tentative de fermeture des votes déjà fermés")
                return JsonResponse({'success': False, 'message': 'Les votes sont déjà fermés!'})
            
            config.votes_ouverts = False
            config.save()
            logger.info("Votes fermés avec succès")
            return JsonResponse({'success': True, 'message': 'Votes fermés avec succès!'})
        
        else:
            logger.warning(f"Action non reconnue: {action}")
            return JsonResponse({'success': False, 'message': 'Action non valide'}, status=400)
    
    except Exception as e:
        logger.error(f"Erreur critique dans controler_periode: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Erreur serveur: {str(e)}'}, status=500)


################MATRICULES ADMINS##################


@require_POST
@admin_required
def generer_matricules_admin(request):
    quantite = int(request.POST.get('quantite', 0))
    created = []

    for _ in range(quantite):
        matricule = "ADM" + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        if not Admin.objects.filter(matricule=matricule).exists():
            admin = Admin.objects.create(matricule=matricule)
            created.append(matricule)

    return JsonResponse({'success': True, 'quantite': len(created), 'codes': created})

@admin_required
def liste_admins(request):
    admins = Admin.objects.filter(est_actif=True).values('matricule')
    return JsonResponse({
        'success': True,
        'admins': list(admins)
    })



@require_POST
@admin_required
def supprimer_admin(request):
    try:
        matricule = request.POST.get('matricule')
        if not matricule:
            return JsonResponse({'success': False, 'message': 'Matricule manquant'}, status=400)
            
        # Suppression de l'admin
        admin = Admin.objects.get(matricule=matricule)
        admin.est_actif = False  # Désactivation plutôt que suppression
        admin.save()
        
        return JsonResponse({'success': True, 'message': 'Admin désactivé avec succès'})
    except Admin.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Admin non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


################VOTE RESPONSABLES CLASSES#############3


@matricule_required
def vote_responsables_classe(request):
    voter = request.voter
    
    # Récupère les postes spécifiques
    postes = ResponsableClassePoste.objects.all()
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return handle_ajax_request(request, voter, postes)
        return handle_regular_post(request, voter, postes)
    
    # GET: Préparer les candidats par poste
    candidats_par_poste = {
        poste: poste.candidats.filter(classe=voter.classe)
        for poste in postes
    }
    
    return render(request, 'vote/responsables_classe.html', {
        'postes': postes,
        'candidats_par_poste': candidats_par_poste,
        'classe_eleve': voter.classe
    })

def handle_ajax_request(request, voter, postes):
    """Gère les requêtes AJAX pour l'enregistrement des votes"""
    # Vérifier si déjà voté
    if VoteResponsableClasse.objects.filter(voter=voter).exists():
        return JsonResponse({
            'success': False,
            'message': 'Vous avez déjà voté pour les responsables de classe.'
        }, status=400)
    
    # Valider chaque vote
    selected_postes = []
    for poste in postes:
        candidat_id = request.POST.get(f'poste_{poste.id}')
        if not candidat_id:
            return JsonResponse({
                'success': False,
                'message': 'Vote incomplet. Vous devez sélectionner un candidat pour chaque poste.'
            }, status=400)
        
        try:
            candidat = ResponsableClasseCandidat.objects.get(
                id=candidat_id,
                poste=poste,
                classe=voter.classe  # Sécurité supplémentaire
            )
            selected_postes.append(poste.id)
        except (ValueError, ResponsableClasseCandidat.DoesNotExist):
            return JsonResponse({
                'success': False,
                'message': 'Candidat invalide détecté.'
            }, status=400)
    
    # Vérifier que tous les postes ont été sélectionnés
    if len(selected_postes) != postes.count():
        return JsonResponse({
            'success': False,
            'message': 'Vote incomplet. Vous devez sélectionner un candidat pour chaque poste.'
        }, status=400)
    
    # Enregistrer les votes
    for poste in postes:
        candidat_id = request.POST.get(f'poste_{poste.id}')
        candidat = ResponsableClasseCandidat.objects.get(id=candidat_id)
        VoteResponsableClasse.objects.create(
            voter=voter,
            candidat=candidat,
            poste=poste
        )
    
    return JsonResponse({'success': True})

def handle_regular_post(request, voter, postes):
    """Gère les soumissions de formulaire traditionnelles (fallback)"""
    if VoteResponsableClasse.objects.filter(voter=voter).exists():
        messages.error(request, "Vous avez déjà voté pour les responsables de classe.")
        return redirect('vote_responsables_classe')
    
    votes_valides = True
    
    # Valider chaque vote
    for poste in postes:
        candidat_id = request.POST.get(f'poste_{poste.id}')
        try:
            candidat = ResponsableClasseCandidat.objects.get(
                id=candidat_id,
                poste=poste,
                classe=voter.classe
            )
        except (ValueError, ResponsableClasseCandidat.DoesNotExist):
            votes_valides = False
            break
    
    if votes_valides:
        # Enregistrer les votes
        for poste in postes:
            candidat_id = request.POST.get(f'poste_{poste.id}')
            candidat = ResponsableClasseCandidat.objects.get(id=candidat_id)
            VoteResponsableClasse.objects.create(
                voter=voter,
                candidat=candidat,
                poste=poste
            )
        
        messages.success(request, "Merci pour votre vote !")
        return redirect('merci')
    
    messages.error(request, "Vote invalide détecté.")
    return redirect('vote_responsables_classe')


###############RESULTATS RESPONSABLES CLASSES#############3



def results_votes_classe(request):
    if 'voter_id' not in request.session:
        return redirect('login')

    try:
        voter = Voter.objects.get(id=request.session['voter_id'])
    except Voter.DoesNotExist:
        return redirect('login')

    # Récupération des données
    postes = ResponsableClassePoste.objects.all().order_by('ordre')
    classes = Voter.CLASS_CHOICES
    
    # Préparation des données de participation
    participation_data = {}
    for classe_code, classe_nom in classes:
        total_voters = Voter.objects.filter(classe=classe_code).count()
        voted_count = VoteResponsableClasse.objects.filter(
            voter__classe=classe_code
        ).values('voter').distinct().count()
        
        participation_data[classe_code] = {
            'classe_nom': classe_nom,
            'total_voters': total_voters,
            'voters_count': voted_count,
            'participation': round((voted_count / total_voters * 100), 1) if total_voters > 0 else 0
        }

    # Structure finale des résultats
    resultats = []
    
    for poste in postes:
        poste_data = {
            'poste': poste,
            'classes': []
        }
        
        for classe_code, classe_nom in classes:
            # Récupération des candidats et votes pour ce poste/classe
            candidats = ResponsableClasseCandidat.objects.filter(
                poste=poste,
                classe=classe_code
            ).annotate(
                vote_count=Count('voteresponsableclasse')
            ).order_by('-vote_count')
            
            total_votes = VoteResponsableClasse.objects.filter(
                poste=poste,
                candidat__classe=classe_code
            ).count()
            
            if candidats.exists():
                gagnant = candidats.first()
                percentage = round((gagnant.vote_count / total_votes * 100), 1) if total_votes > 0 else 0
                
                poste_data['classes'].append({
                    'classe_code': classe_code,
                    'classe_nom': classe_nom,
                    'gagnant': {
                        'id': gagnant.id,
                        'nom': gagnant.nom,
                        'photo': gagnant.photo,
                        'vote_count': gagnant.vote_count,
                        'percentage': percentage
                    },
                    'total_votes': total_votes,
                    'vote_count': gagnant.vote_count,
                    'percentage': percentage,
                    'candidats': [
                        {
                            'nom': c.nom,
                            'photo': c.photo,
                            'vote_count': c.vote_count,
                            'percentage': round((c.vote_count / total_votes * 100), 1) if total_votes > 0 else 0
                        }
                        for c in candidats
                    ],
                    **participation_data.get(classe_code, {})
                })
            else:
                poste_data['classes'].append({
                    'classe_code': classe_code,
                    'classe_nom': classe_nom,
                    'gagnant': None,
                    'total_votes': total_votes,
                    'vote_count': 0,
                    'percentage': 0,
                    'candidats': [],
                    **participation_data.get(classe_code, {})
                })
        
        resultats.append(poste_data)

    # Correction de la préparation des données pour la section participation
    participation_stats = [
        {
            'classe_code': code,
            'classe_nom': nom,
            'voters_count': participation_data[code]['voters_count'],
            'total_voters': participation_data[code]['total_voters'],
            'participation': participation_data[code]['participation']
        }
        for code, nom in classes  # Modification ici - itération directe sur les tuples
    ]

    context = {
        'resultats': resultats,
        'participation_stats': participation_stats,
        'classe_eleve': voter.classe
    }
    
    return render(request, 'vote/results_votes_classe.html', context)


def results_portal(request):
    return render(request, 'vote/results_prime.html')



###############PRESIDENTS CLUBS#############3



@matricule_required
def vote_club(request):
    clubs = Club.objects.prefetch_related('candidats').all()
    
    # Vérifier si l'utilisateur a déjà voté pour chaque club
    voted_clubs = set()
    if hasattr(request, 'voter'):
        voted_clubs = set(VoteClub.objects.filter(voter=request.voter)
                         .values_list('club_id', flat=True))
    
    # Préparer les données des clubs avec les candidats formatés
    clubs_data = []
    for club in clubs:
        candidats_data = []
        for candidat in club.candidats.all():
            # Formatage du nom avec la classe si elle existe
            nom_complet = candidat.nom
            if candidat.classe:
                nom_complet += f" ({candidat.classe})"
            
            candidats_data.append({
                'id': candidat.id,
                'nom_complet': nom_complet,
                'slogan': candidat.slogan,
                'photo': candidat.photo,
                'programme': candidat.programme
            })
        
        clubs_data.append({
            'id': club.id,
            'nom': club.nom,
            'description': club.description,
            'logo': club.logo,
            'candidats': candidats_data
        })
    
    return render(request, 'vote/vote_club.html', {
        'clubs': clubs_data,
        'voted_clubs': voted_clubs
    })

@require_POST
@matricule_required
def vote_club_submit(request):
    print("Requête reçue pour vote_club_submit")  # Debug
    try:
        data = json.loads(request.body)
        print("Données reçues:", data)  # Debug
        
        candidate_id = data.get('candidate_id')
        club_id = data.get('club_id')
        
        # Validation des données
        if not candidate_id or not club_id:
            print("Données manquantes")  # Debug
            return JsonResponse({'success': False, 'message': 'Données manquantes'}, status=400)
        
        # Vérifier si l'utilisateur a déjà voté
        if VoteClub.objects.filter(voter=request.voter, club_id=club_id).exists():
            print("Déjà voté")  # Debug
            return JsonResponse({'success': False, 'message': 'Vous avez déjà voté pour ce club'}, status=400)
        
        # Vérifier que le candidat appartient au club
        if not ClubCandidat.objects.filter(id=candidate_id, club_id=club_id).exists():
            print("Candidat invalide")  # Debug
            return JsonResponse({'success': False, 'message': 'Candidat invalide'}, status=400)
        
        # Enregistrer le vote
        vote = VoteClub.objects.create(
            voter=request.voter,
            candidat_id=candidate_id,
            club_id=club_id
        )
        print(f"Vote enregistré ID: {vote.id}")  # Debug
        
        return JsonResponse({'success': True, 'message': 'Vote enregistré'})
        
    except Exception as e:
        print("Erreur:", str(e))  # Debug
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@matricule_required
def club_stats(request):
    # Logique pour les statistiques
    clubs = Club.objects.all()
    stats = []
    
    for club in clubs:
        total_votes = VoteClub.objects.filter(club=club).count()
        candidates = []
        
        for candidate in club.candidats.all():
            vote_count = VoteClub.objects.filter(candidat=candidate).count()
            percentage = (vote_count / total_votes * 100) if total_votes > 0 else 0
            
            candidates.append({
                'name': candidate.nom,
                'photo': candidate.photo.url if candidate.photo else None,
                'vote_count': vote_count,
                'percentage': round(percentage, 1)
            })
        
        # Trier les candidats par nombre de votes décroissant
        candidates.sort(key=lambda x: x['vote_count'], reverse=True)
        
        stats.append({
            'club': club,
            'total_votes': total_votes,
            'candidates': candidates,
            'winner': candidates[0] if candidates else None
        })
    
    return render(request, 'vote/club_stats.html', {
        'stats': stats
    })


from datetime import datetime
from django.db.models import Prefetch, Count
import json
from django.shortcuts import render
from .models import Club, ClubCandidat, VoteClub, Voter

def results_club_view(request):
    current_year = datetime.now().year
    
    # Récupérer tous les clubs avec leurs candidats et votes
    clubs = Club.objects.prefetch_related(
        Prefetch('candidats', 
                queryset=ClubCandidat.objects.annotate(
                    vote_count=Count('voteclub')
                )
        )
    ).annotate(
        total_votes=Count('candidats__voteclub')
    ).order_by('-total_votes')
    
    # Préparer les données pour le frontend
    clubs_data = []
    for club in clubs:
        candidates = []
        for candidate in club.candidats.all():
            candidates.append({
                'id': candidate.id,
                'nom': candidate.nom,
                'vote_count': candidate.vote_count,
                'photo_url': candidate.photo.url if candidate.photo else '/static/images/default_avatar.png',
                'club_nom': club.nom
            })
        
        clubs_data.append({
            'id': club.id,
            'nom': club.nom,
            'total_votes': club.total_votes,
            'candidates': sorted(candidates, key=lambda x: x['vote_count'], reverse=True)
        })
    
    # Préparer les données JSON
    context = {
        'current_year': current_year,
        'clubs': clubs,
        'CLASS_CHOICES': Voter.CLASS_CHOICES,
        'clubs_json': json.dumps(clubs_data),
        'participation_json': json.dumps({
            'byFiliere': {
                'ISE': {
                    'voters': VoteClub.objects.filter(voter__filiere='ISE').count(),
                    'total': Voter.objects.filter(filiere='ISE').count(),
                    'percentage': round(VoteClub.objects.filter(voter__filiere='ISE').count() / 
                             Voter.objects.filter(filiere='ISE').count() * 100, 1) if Voter.objects.filter(filiere='ISE').count() > 0 else 0
                },
                'AS': {
                    'voters': VoteClub.objects.filter(voter__filiere='AS').count(),
                    'total': Voter.objects.filter(filiere='AS').count(),
                    'percentage': round(VoteClub.objects.filter(voter__filiere='AS').count() / 
                             Voter.objects.filter(filiere='AS').count() * 100, 1) if Voter.objects.filter(filiere='AS').count() > 0 else 0
                }
            },
            'global': {
                'voters': VoteClub.objects.count(),
                'total': Voter.objects.count(),
                'percentage': round(VoteClub.objects.count() / Voter.objects.count() * 100, 1) if Voter.objects.count() > 0 else 0
            }
        }),
        'timeline_json': json.dumps({
            'labels': [f"{hour}:00" for hour in range(8, 21, 2)],
            'data': [
                VoteClub.objects.filter(
                    date_vote__hour__lt=hour + 2,
                    date_vote__hour__gte=hour
                ).count() for hour in range(8, 21, 2)
            ]
        })
    }
    
    return render(request, 'vote/results_club.html', context)



###############CANDIDATURES PRESIDENTS CLUBS#############3



@matricule_required
def candidature_club_view(request):
    voter = request.voter  # injecté par le décorateur

    # Vérifier si l'utilisateur est déjà candidat
    # On cherche si un candidat existe avec la même classe et un nom qui pourrait correspondre
    deja_candidat = ClubCandidat.objects.filter(
        nom__endswith=f"({voter.classe})"
    ).exists()
    
    if deja_candidat:
        messages.warning(request, "Vous avez déjà soumis une candidature.")
        return redirect('accueil')

    if request.method == 'POST':
        form = ClubCandidatureForm(request.POST, request.FILES)
        if form.is_valid():
            club = form.cleaned_data['club']
            prenom = form.cleaned_data['prenom'].strip().title()  # Capitalize first letter
            nom = form.cleaned_data['nom'].strip().upper()  # Force le nom en majuscules
            photo = form.cleaned_data['photo']
            slogan = form.cleaned_data['slogan']
            programme = form.cleaned_data['programme']

            # Format: "Prénom NOM (Classe)"
            nom_complet = f"{prenom} {nom}"

            # Création du candidat
            ClubCandidat.objects.create(
                nom=nom_complet,
                club=club,
                photo=photo,
                slogan=slogan,
                programme=programme,
                classe=voter.classe,
                filiere=voter.filiere
            )
            
            messages.success(request, f"Votre candidature pour {club.nom} a été enregistrée avec succès !")
            return redirect('accueil')
    else:
        form = ClubCandidatureForm()

    return render(request, 'vote/candidature_club.html', {
        'form': form,
        'deja_candidat': deja_candidat,
        'classe_actuelle': voter.classe
    })



@matricule_required
def candidature_responsable_view(request):
    voter = request.voter

    # Vérifier si l'utilisateur est déjà candidat
    deja_candidat = ResponsableClasseCandidat.objects.filter(
        matricule=voter.matricule
    ).exists()
    
    if deja_candidat:
        messages.warning(request, "Vous avez déjà soumis une candidature pour un poste de responsable.")
        return redirect('accueil')

    if request.method == 'POST':
        form = ResponsableClasseCandidatureForm(request.POST, request.FILES)
        if form.is_valid():
            poste = form.cleaned_data['poste']
            prenom = form.cleaned_data['prenom'].strip().title()
            nom = form.cleaned_data['nom'].strip().upper()
            photo = form.cleaned_data['photo']
            programme = form.cleaned_data['programme']

            nom_complet = f"{prenom} {nom}"

            ResponsableClasseCandidat.objects.create(
                nom=nom_complet,
                poste=poste,
                photo=photo,
                programme=programme,  # Champ maintenant disponible
                classe=voter.classe,
                filiere=voter.filiere,
                matricule=voter.matricule
            )
            
            messages.success(request, f"Votre candidature pour {poste.nom} a été enregistrée !")
            return redirect('accueil')
    else:
        form = ResponsableClasseCandidatureForm()

    return render(request, 'vote/candidature_responsable.html', {
        'form': form,
        'deja_candidat': deja_candidat,
        'classe_actuelle': voter.classe
    })


def portail_candidatures(request):
    config = Configuration.get_config()
    if not config.candidatures_ouvertes:
        return redirect('access_closed')  # Ou une autre page d'erreur

    return render(request, 'vote/candidature_portail.html')



@check_access('aes_elections')
def inscription_aes(request):
    # Votre vue normale
    return render(request, 'candidature_portail.html')

def access_closed(request):
    return render(request, 'vote/access_closed.html')

def access_closed_vote(request):
    from django.template.loader import get_template
    try:
        template = get_template('vote/access_closed_vote.html')
        print("Template trouvé à:", template.origin.name)
        return render(request, 'vote/access_closed_vote.html')
    except Exception as e:
        print("Erreur template:", str(e))
        return HttpResponse("Template error - check console", status=500)
    

#############OBTENTION MATRICULES#########


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Max
import pandas as pd
import logging

from .models import EmailAutorise, Voter

logger = logging.getLogger(__name__)


@admin_required
def obtenir_matricule_view(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        if not email:
            messages.error(request, "Veuillez saisir une adresse e-mail.")
            return render(request, 'vote/demande_matricule.html')

        try:
            email_ref = EmailAutorise.objects.get(email=email)

            if email_ref.matricule_attribue:
                return render(request, 'vote/recuperation_matricule.html', {
                    'email': email,
                    'matricule': email_ref.matricule_attribue,
                    'deja_attribue': True
                })

            voter_libre = Voter.objects.filter(
                classe=email_ref.classe,
                date_attribution__isnull=True
            ).order_by('?').first()

            if not voter_libre:
                messages.error(request, "Aucun matricule disponible pour votre classe.")
                return render(request, 'vote/demande_matricule.html')

            voter_libre.date_attribution = timezone.now()
            voter_libre.save()

            email_ref.matricule_attribue = voter_libre.matricule
            email_ref.save()

            return render(request, 'vote/recuperation_matricule.html', {
                'email': email,
                'matricule': voter_libre.matricule,
                'deja_attribue': False
            })

        except EmailAutorise.DoesNotExist:
            messages.error(request, "Cet email n’est pas autorisé.")
            return render(request, 'vote/demande_matricule.html')

    return render(request, 'vote/demande_matricule.html')



@admin_required
def importer_emails_autorises_view(request):
    stats = {
        'total_emails': EmailAutorise.objects.count(),
        'dernier_import': EmailAutorise.objects.aggregate(
            dernier=Max('id')
        )['dernier']
    }

    if request.method == 'POST' and 'fichier_excel' in request.FILES:
        fichier = request.FILES['fichier_excel']
        try:
            df = pd.read_excel(fichier)
            compteur = 0
            erreurs = 0

            for index, row in df.iterrows():
                try:
                    email = str(row.get('email', '')).strip().lower()
                    classe = str(row.get('classe', '')).strip()

                    if not email or not classe:
                        continue

                    EmailAutorise.objects.update_or_create(
                        email=email,
                        defaults={'classe': classe}
                    )
                    compteur += 1
                except Exception as e:
                    erreurs += 1
                    logger.error(f"Erreur ligne {index + 2}: {str(e)}")

            messages.success(request, f"Import terminé : {compteur} emails | {erreurs} erreurs")
            return redirect('importer_emails_autorises')

        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier : {str(e)}")

    emails = EmailAutorise.objects.all().order_by('classe', 'email')
    return render(request, 'vote/importer_emails_autorises.html', {
        'total_emails': stats['total_emails'],
        'dernier_import': stats['dernier_import'],
        'emails': emails
    })




@admin_required
def reinitialiser_matricule(request, email_id):
    email = get_object_or_404(EmailAutorise, id=email_id)
    if email.matricule_attribue:
        Voter.objects.filter(matricule=email.matricule_attribue).update(date_attribution=None)
        email.matricule_attribue = None
        email.save()
        messages.success(request, f"Matricule réinitialisé pour {email.email}.")
    else:
        messages.info(request, f"Aucun matricule n'était attribué à {email.email}.")
    return redirect('importer_emails_autorises')




@admin_required
def supprimer_email_autorise(request, email_id):
    email = get_object_or_404(EmailAutorise, id=email_id)
    if email.matricule_attribue:
        Voter.objects.filter(matricule=email.matricule_attribue).delete()
    email.delete()
    messages.success(request, f"Email {email.email} supprimé avec succès.")
    return redirect('importer_emails_autorises')




################GENERATION DU RAPPORT######################


from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # ⛔ Ne pas utiliser GUI
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

from .models import Poste, Candidat, VoteParPoste, Voter

def generate_pdf_report(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    elements = []
    image_buffers = []

    try:
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, spaceAfter=20,
                                     textColor=colors.HexColor('#2c3e50'), fontName='Helvetica-Bold')
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], alignment=1, spaceAfter=10)
        section_style = ParagraphStyle('Section', parent=styles['Heading3'], spaceBefore=10, spaceAfter=10)

        # Titre
        elements.append(Paragraph("RAPPORT DÉTAILLÉ DES RÉSULTATS ÉLECTORAUX", title_style))
        elements.append(Paragraph(f"Élections de l'AES - Année {datetime.now().year}", subtitle_style))
        elements.append(Spacer(1, 0.5 * inch))

        # Section Participation par Filière
        elements.append(Paragraph("PARTICIPATION PAR FILIÈRE", section_style))
        filieres = ['ISE', 'AS']
        participation_data = []

        for filiere in filieres:
            total = Voter.objects.filter(filiere=filiere).count()
            votes = VoteParPoste.objects.filter(voter__filiere=filiere).values('voter').distinct().count()
            participation_data.append({
                'Filière': filiere,
                'Inscrits': total,
                'Votants': votes,
                'Taux': (votes / total * 100) if total > 0 else 0
            })

        participation_df = pd.DataFrame(participation_data)
        plt.figure(figsize=(8, 4))
        sns.set_theme(style="whitegrid")
        ax = sns.barplot(x='Filière', y='Taux', data=participation_df, palette="Blues_d")
        ax.set_title('Taux de participation par filière', fontsize=14)
        ax.set_ylabel('Taux de participation (%)')
        ax.set_ylim(0, 100)

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=300)
        plt.close()
        img_buffer.seek(0)
        image_buffers.append(img_buffer)
        elements.append(Image(img_buffer, width=6 * inch, height=3 * inch))
        elements.append(Spacer(1, 0.3 * inch))

        # Résultats par Poste
        elements.append(Paragraph("RÉSULTATS DÉTAILLÉS PAR POSTE", section_style))
        postes = Poste.objects.all().order_by('code')

        for poste in postes:
            candidats = Candidat.objects.filter(poste=poste).annotate(
                vote_count=models.Count('voteparposte')
            ).order_by('-vote_count')

            if not candidats.exists():
                continue

            data = [['Nom', 'Classe', 'Votes']]
            for c in candidats:
                classe_label = getattr(c.classe, 'nom', c.classe)
                data.append([c.nom, classe_label if classe_label else '-', c.vote_count])


            table = Table(data, hAlign='LEFT', colWidths=[3 * inch, 2 * inch, 1 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(Paragraph(poste.nom, styles['Heading4']))
            elements.append(table)

            # Graphique
            chart_data = [{'Candidat': c.nom[:15] + '...' if len(c.nom) > 15 else c.nom, 'Votes': c.vote_count}
                          for c in candidats]
            chart_df = pd.DataFrame(chart_data)

            plt.figure(figsize=(8, 4))
            sns.set_palette("husl")
            ax = sns.barplot(x='Candidat', y='Votes', data=chart_df)
            ax.set_title(f'Répartition des votes - {poste.nom}', fontsize=12)
            ax.set_ylabel('Nombre de votes')
            ax.set_xlabel('Candidats')
            plt.xticks(rotation=45, ha='right')

            poste_buffer = BytesIO()
            plt.savefig(poste_buffer, format='png', bbox_inches='tight', dpi=300)
            plt.close()
            poste_buffer.seek(0)
            image_buffers.append(poste_buffer)
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Image(poste_buffer, width=6 * inch, height=3 * inch))
            elements.append(Spacer(1, 0.5 * inch))

        # Générer le document PDF
        doc.build(elements)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_elections_aes.pdf"'
        response.write(buffer.getvalue())
        return response

    finally:
        buffer.close()
        for buf in image_buffers:
            buf.close()
        plt.close('all')  # Fermer toutes les figures résiduelles
